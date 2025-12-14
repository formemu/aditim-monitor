"""
Окно отчётов.

Предоставляет интерфейс для формирования различных отчётов по производственным данным.
"""

from datetime import datetime, date
from typing import Optional, List, Dict, Any
from PySide6.QtCore import QFile, Qt, QDate
from PySide6.QtUiTools import QUiLoader
from PySide6.QtWidgets import QWidget, QTableWidgetItem, QFileDialog, QMessageBox

from ..base_table import BaseTable
from ..constant import UI_PATHS_ABS as UI_PATHS
from ..api_manager import api_manager


class WindowReport(QWidget):
    """
    Окно отчётов.
    
    Функционал:
    - Выбор типа отчёта
    - Настройка периода (месяц или диапазон)
    - Фильтрация данных
    - Формирование отчётов
    - Экспорт в Excel
    """

    def __init__(self):
        """Инициализация окна отчётов."""
        super().__init__()
        self.api_manager = api_manager
        self.current_report_data = []
        
        self.load_ui()
        self.setup_ui()
        self.connect_signals()

    def load_ui(self) -> None:
        """Загрузка UI из .ui файла."""
        ui_file = QFile(UI_PATHS["CONTENT_REPORT"])
        ui_file.open(QFile.ReadOnly)
        
        loader = QUiLoader()
        self.ui = loader.load(ui_file)
        ui_file.close()

    def setup_ui(self) -> None:
        """Настройка UI компонентов."""
        # Устанавливаем текущий год в комбобокс
        current_year = datetime.now().year
        year_index = self.ui.comboBox_year.findText(str(current_year))
        if year_index >= 0:
            self.ui.comboBox_year.setCurrentIndex(year_index)
        
        # Устанавливаем текущий месяц
        current_month = datetime.now().month - 1  # Индексы с 0
        self.ui.comboBox_month.setCurrentIndex(current_month)
        
        # Устанавливаем даты по умолчанию
        today = QDate.currentDate()
        self.ui.dateEdit_from.setDate(today.addMonths(-1))
        self.ui.dateEdit_to.setDate(today)
        
        # Загружаем справочники для фильтров
        self.load_directories()

    def load_directories(self) -> None:
        """Загрузка справочников в фильтры."""
        # Отделы
        self.ui.comboBox_department.clear()
        for item in self.api_manager.directory.get("dir_department", []):
            self.ui.comboBox_department.addItem(
                item["description"], 
                item["id"]
            )
        
        # Типы задач
        self.ui.comboBox_task_type.clear()
        for item in self.api_manager.directory.get("dir_task_type", []):
            self.ui.comboBox_task_type.addItem(
                item["description"], 
                item["id"]
            )
        
        # Статусы задач
        self.ui.comboBox_status.clear()
        for item in self.api_manager.directory.get("dir_task_status", []):
            self.ui.comboBox_status.addItem(
                item["description"], 
                item["id"]
            )

    def connect_signals(self) -> None:
        """Подключение сигналов к слотам."""
        # Переключение режима периода
        self.ui.radioButton_month.toggled.connect(self.on_period_mode_changed)
        self.ui.radioButton_range.toggled.connect(self.on_period_mode_changed)
        
        # Фильтры
        self.ui.checkBox_filter_department.toggled.connect(
            lambda checked: self.ui.comboBox_department.setEnabled(checked)
        )
        self.ui.checkBox_filter_task_type.toggled.connect(
            lambda checked: self.ui.comboBox_task_type.setEnabled(checked)
        )
        self.ui.checkBox_filter_status.toggled.connect(
            lambda checked: self.ui.comboBox_status.setEnabled(checked)
        )
        
        # Кнопки
        self.ui.pushButton_generate.clicked.connect(self.on_generate_report)
        self.ui.pushButton_export.clicked.connect(self.on_export_excel)

    def on_period_mode_changed(self, checked: bool) -> None:
        """
        Обработка переключения режима выбора периода.

        Args:
            checked: Состояние радиокнопки
        """
        is_month_mode = self.ui.radioButton_month.isChecked()
        
        self.ui.comboBox_month.setEnabled(is_month_mode)
        self.ui.comboBox_year.setEnabled(is_month_mode)
        self.ui.widget_date_range.setEnabled(not is_month_mode)

    def get_date_range(self) -> tuple[date, date]:
        """
        Получение диапазона дат для отчёта.

        Returns:
            Кортеж (date_from, date_to)
        """
        if self.ui.radioButton_month.isChecked():
            # Режим месяца
            year = int(self.ui.comboBox_year.currentText())
            month = self.ui.comboBox_month.currentIndex() + 1
            
            # Первый день месяца
            date_from = date(year, month, 1)
            
            # Последний день месяца
            if month == 12:
                date_to = date(year, 12, 31)
            else:
                date_to = date(year, month + 1, 1)
                date_to = date(date_to.year, date_to.month, date_to.day - 1)
        else:
            # Режим произвольного диапазона
            date_from = self.ui.dateEdit_from.date().toPython()
            date_to = self.ui.dateEdit_to.date().toPython()
        
        return date_from, date_to

    def get_filters(self) -> Dict[str, Any]:
        """
        Получение активных фильтров.

        Returns:
            Словарь с активными фильтрами
        """
        filters = {}
        
        if self.ui.checkBox_filter_department.isChecked():
            filters["department_id"] = self.ui.comboBox_department.currentData()
        
        if self.ui.checkBox_filter_task_type.isChecked():
            filters["task_type_id"] = self.ui.comboBox_task_type.currentData()
        
        if self.ui.checkBox_filter_status.isChecked():
            filters["task_status_id"] = self.ui.comboBox_status.currentData()
        
        return filters

    def on_generate_report(self) -> None:
        """Обработка формирования отчёта."""
        report_type = self.ui.comboBox_report_type.currentIndex()
        
        if report_type == 0:
            self.generate_report_task()
        elif report_type == 1:
            self.generate_report_machine()
        elif report_type == 2:
            self.generate_report_blank()
        elif report_type == 3:
            self.generate_report_profiletool()
        elif report_type == 4:
            self.generate_report_summary()
        
        # Активируем кнопку экспорта
        self.ui.pushButton_export.setEnabled(True)

    def generate_report_task(self) -> None:
        """Формирование отчёта по задачам."""
        date_from, date_to = self.get_date_range()
        filters = self.get_filters()
        
        # Получаем все задачи
        list_task = self.api_manager.table.get("task", [])
        
        # Фильтруем по периоду
        list_filtered = []
        for task in list_task:
            task_date = datetime.fromisoformat(task["created"]).date()
            
            if date_from <= task_date <= date_to:
                # Применяем дополнительные фильтры
                if "department_id" in filters:
                    if task.get("department_id") != filters["department_id"]:
                        continue
                
                if "task_type_id" in filters:
                    if task.get("task_type_id") != filters["task_type_id"]:
                        continue
                
                if "task_status_id" in filters:
                    if task.get("task_status_id") != filters["task_status_id"]:
                        continue
                
                list_filtered.append(task)
        
        # Сохраняем данные для экспорта
        self.current_report_data = list_filtered
        
        # Формируем таблицу
        list_header = [
            "Номер",
            "Тип",
            "Отдел",
            "Статус",
            "Местоположение",
            "Создана",
            "Срок",
            "Описание"
        ]
        
        def map_row(task: Dict[str, Any]) -> List[str]:
            task_type = self.api_manager.get_by_id("dir_task_type", task.get("task_type_id"))
            department = self.api_manager.get_by_id("dir_department", task.get("department_id"))
            status = self.api_manager.get_by_id("dir_task_status", task.get("task_status_id"))
            location = self.api_manager.get_by_id("dir_task_location", task.get("task_location_id"))
            
            created = datetime.fromisoformat(task["created"]).strftime("%d.%m.%Y")
            deadline = datetime.fromisoformat(task["deadline"]).strftime("%d.%m.%Y") if task.get("deadline") else "-"
            
            return [
                task.get("number", ""),
                task_type.get("description", "") if task_type else "",
                department.get("description", "") if department else "",
                status.get("description", "") if status else "",
                location.get("description", "") if location else "",
                created,
                deadline,
                task.get("description", "")
            ]
        
        BaseTable.populate_table(
            self.ui.tableWidget_report,
            list_header,
            list_filtered,
            func_row_mapper=map_row,
            func_id_getter=lambda t: t["id"]
        )
        
        # Формируем сводку
        self.generate_summary_task(list_filtered, date_from, date_to)

    def generate_summary_task(self, list_task: List[Dict[str, Any]], date_from: date, date_to: date) -> None:
        """
        Формирование сводки по задачам.

        Args:
            list_task: Список задач
            date_from: Начало периода
            date_to: Конец периода
        """
        total = len(list_task)
        
        # Группировка по статусам
        dict_status = {}
        for task in list_task:
            status_id = task.get("task_status_id")
            status = self.api_manager.get_by_id("dir_task_status", status_id)
            status_name = status.get("description", "Неизвестно") if status else "Неизвестно"
            
            dict_status[status_name] = dict_status.get(status_name, 0) + 1
        
        # Группировка по отделам
        dict_department = {}
        for task in list_task:
            dept_id = task.get("department_id")
            dept = self.api_manager.get_by_id("dir_department", dept_id)
            dept_name = dept.get("description", "Неизвестно") if dept else "Неизвестно"
            
            dict_department[dept_name] = dict_department.get(dept_name, 0) + 1
        
        # Формируем текст сводки
        summary = f"""
СВОДКА ПО ЗАДАЧАМ
Период: {date_from.strftime('%d.%m.%Y')} - {date_to.strftime('%d.%m.%Y')}

═══════════════════════════════════════════════════════

ОБЩАЯ СТАТИСТИКА:
  • Всего задач: {total}

ПО СТАТУСАМ:
"""
        
        for status_name, count in sorted(dict_status.items()):
            percent = (count / total * 100) if total > 0 else 0
            summary += f"  • {status_name}: {count} ({percent:.1f}%)\n"
        
        summary += "\nПО ОТДЕЛАМ:\n"
        for dept_name, count in sorted(dict_department.items()):
            percent = (count / total * 100) if total > 0 else 0
            summary += f"  • {dept_name}: {count} ({percent:.1f}%)\n"
        
        self.ui.textEdit_summary.setPlainText(summary)

    def generate_report_machine(self) -> None:
        """Формирование отчёта по станкам."""
        QMessageBox.information(
            self.ui,
            "В разработке",
            "Отчёт по станкам будет реализован в следующей версии."
        )

    def generate_report_blank(self) -> None:
        """Формирование отчёта по заготовкам."""
        QMessageBox.information(
            self.ui,
            "В разработке",
            "Отчёт по заготовкам будет реализован в следующей версии."
        )

    def generate_report_profiletool(self) -> None:
        """Формирование отчёта по инструментам."""
        QMessageBox.information(
            self.ui,
            "В разработке",
            "Отчёт по инструментам будет реализован в следующей версии."
        )

    def generate_report_summary(self) -> None:
        """Формирование сводного отчёта."""
        QMessageBox.information(
            self.ui,
            "В разработке",
            "Сводный отчёт будет реализован в следующей версии."
        )

    def on_export_excel(self) -> None:
        """Экспорт отчёта в Excel."""
        if not self.current_report_data:
            QMessageBox.warning(
                self.ui,
                "Нет данных",
                "Сначала сформируйте отчёт."
            )
            return
        
        # Диалог сохранения файла
        file_path, _ = QFileDialog.getSaveFileName(
            self.ui,
            "Экспорт в Excel",
            f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            self.export_to_excel(file_path)
            QMessageBox.information(
                self.ui,
                "Успех",
                f"Отчёт успешно экспортирован в:\n{file_path}"
            )
        except Exception as e:
            QMessageBox.critical(
                self.ui,
                "Ошибка экспорта",
                f"Не удалось экспортировать отчёт:\n{str(e)}"
            )

    def export_to_excel(self, file_path: str) -> None:
        """
        Экспорт данных в Excel файл.

        Args:
            file_path: Путь для сохранения файла
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            QMessageBox.warning(
                self.ui,
                "Библиотека не найдена",
                "Для экспорта в Excel необходимо установить библиотеку openpyxl:\n"
                "pip install openpyxl"
            )
            return
        
        # Создаём книгу
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Отчёт"
        
        # Заголовок
        table = self.ui.tableWidget_report
        
        # Записываем заголовки
        for col in range(table.columnCount()):
            header_item = table.horizontalHeaderItem(col)
            cell = ws.cell(row=1, column=col + 1)
            cell.value = header_item.text() if header_item else ""
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Записываем данные
        for row in range(table.rowCount()):
            for col in range(table.columnCount()):
                item = table.item(row, col)
                cell = ws.cell(row=row + 2, column=col + 1)
                cell.value = item.text() if item else ""
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем
        wb.save(file_path)
